import json, datetime, io, csv
import streamlit as st
import pandas as pd

st.set_page_config(page_title="MIA Stil – Kalkulator Korpusa (Unified V5+)", page_icon="🧮", layout="wide", initial_sidebar_state="expanded")

# =============== Global styles ===============
st.markdown("""
<style>
:root { color-scheme: light only; }
html, body, .stApp { background:#fbfbfc; }
h1,h2,h3 { font-weight:800; letter-spacing:-0.02em; }
.section { background:white; border:1px solid #e5e7eb; border-radius:12px; padding:14px; box-shadow: 0 1px 2px rgba(0,0,0,.03); }
.kv { width:100%; border-collapse:collapse; }
.kv th,.kv td{ border-bottom:1px solid #eef2f7; padding:8px 10px; }
.kv th{ text-align:left; width:55%; background:#f9fafb; }
.kv td{ text-align:right; width:45%; font-variant-numeric: tabular-nums; }
.total{ background:#effbf1 !important; font-weight:800; }
.sticky { position:sticky; top:0; z-index:99; background:rgba(251,251,252,.85); backdrop-filter:blur(6px); border-bottom:1px solid #e5e7eb; padding:8px 0 6px; }
.badge { display:inline-block; padding:2px 8px; border-radius:999px; font-size:.75rem; background:#eef2ff; color:#3730a3; border:1px solid #e0e7ff; }
.help { color:#6b7280; font-size:.9rem; }
.small { font-size:.9rem; color:#374151; }
</style>
""", unsafe_allow_html=True)

# =============== Sidebar: Loader (JSON / upload / CSV URL) ===============
st.sidebar.header("📦 Cjenik – Učitavanje")
src = st.sidebar.radio(
    "Izvor cjenika",
    ["Lokalni cjenik.json (default)", "Učitaj JSON (drag&drop)", "CSV URL-ovi (Google Sheets)"],
    index=0
)

def normalize_cjenik(data: dict):
    """Uskladi sve ključeve i normaliziraj OKOV/OPREMA liste (razne varijante 'art_nr')."""
    data = data or {}

    # Osnovne grupe
    for key, default in [
        ("materijali", []), ("abs_trake", []),
        ("materijali_fronta", []), ("abs_trake_fronta", []),
        ("usluge", []),
        ("okov", []), ("oprema", []), ("dodatci", []),
    ]:
        data.setdefault(key, default)

    # Aliasi (korisnici često napišu množinu ili krivo)
    if not data.get("okov") and data.get("okovi"):
        data["okov"] = data.get("okovi") or []
    if not data.get("oprema") and data.get("opreme"):
        data["oprema"] = data.get("opreme") or []

    # Helper: normaliziraj ključeve jednog retka u katalogu (OKOV/OPREMA)
    def norm_item_keys(x: dict):
        if not isinstance(x, dict):
            return {}
        out = {}
        for k, v in x.items():
            k_norm = str(k).strip().lower().replace(" ", "_").replace("-", "_").replace(".", "")
            # mapiranja za art_nr
            if k_norm in ("art_nr", "artnr", "art__nr", "artnr_", "art_nr_", "artnr__"):
                k_norm = "art_nr"
            if k_norm in ("art", "sifra_artikla", "artikl", "sifra"):  # minimalistički aliasi
                # Samo mapiraj u art_nr ako izgleda kao šifra artikla (string bez razmaka)
                if isinstance(v, str) and v.strip():
                    k_norm = "art_nr"
            out[k_norm] = v
        # standardiziraj tipove/praznine
        if "art_nr" in out and isinstance(out["art_nr"], str):
            out["art_nr"] = out["art_nr"].strip()
        if "jedinica" in out and isinstance(out["jedinica"], str):
            out["jedinica"] = out["jedinica"].strip()
        return out

    # Normaliziraj OKOV/OPREMA zapise (ključeve i whitespace)
    data["okov"] = [norm_item_keys(x) for x in (data.get("okov") or [])]
    data["oprema"] = [norm_item_keys(x) for x in (data.get("oprema") or [])]

    # Ako i dalje nema art_nr, pokušaj iz "art" ili "sifra"
    def ensure_art_nr(lst):
        fixed = []
        for x in lst:
            if not x.get("art_nr"):
                # fallback iz nekoliko mogućih polja
                for alt in ("art", "sifra_artikla", "artikl", "sifra"):
                    val = x.get(alt)
                    if isinstance(val, str) and val.strip():
                        x["art_nr"] = val.strip()
                        break
            fixed.append(x)
        return fixed

    data["okov"] = ensure_art_nr(data["okov"])
    data["oprema"] = ensure_art_nr(data["oprema"])

    # Placeholderi ako je sve prazno
    if not data["okov"]:
        data["okov"] = [{
            "art_nr":"OK-1001", "naziv":"Pant (par) – placeholder",
            "dobavljac":"Blum", "jedinica":"par", "cijena_eur":6.20
        }]
    if not data["oprema"]:
        data["oprema"] = [{
            "art_nr":"OP-2001", "naziv":"Ručkica 160mm – placeholder",
            "dobavljac":"Hettich", "jedinica":"kom", "cijena_eur":3.20
        }]
    if not data.get("dodatci"):
        data["dodatci"] = [{
            "sifra":"DD-001","naziv":"Dodatni element – placeholder",
            "jedinica":"po kom","cijena_eur":10.00,"vrsta":"po kom"
        }]

    return data

@st.cache_data(show_spinner=False)
def load_local():
    with open("cjenik.json","r",encoding="utf-8") as f:
        return normalize_cjenik(json.load(f))

def from_csv_rows(rows, schema):
    out = []
    for r in rows:
        item = {}
        for k, conv in schema.items():
            val = r.get(k, "")
            if conv:
                try:
                    val = conv(val)
                except Exception:
                    val = conv("0")
            item[k] = val
        out.append(item)
    return out

@st.cache_data(show_spinner=False)
def load_from_uploaded(file_bytes: bytes):
    return normalize_cjenik(json.loads(file_bytes.decode("utf-8")))

@st.cache_data(show_spinner=False)
def load_from_csv_urls(url_mat, url_trak, url_fr, url_ftrak, url_usl,
                       url_okov, url_oprema, url_dodatci):
    s_m = {"sifra":str, "naziv":str, "cijena_eur_po_m2": lambda x: float(str(x).replace(",","."))}
    s_t = {"sifra":str, "naziv":str, "cijena_eur_po_m":  lambda x: float(str(x).replace(",","."))}
    s_u = {"sifra":str, "naziv":str, "cijena_eur_po_m":  lambda x: float(str(x).replace(",","."))}
    s_x = {"art_nr":str, "naziv":str, "dobavljac":str, "jedinica":str, "cijena_eur": lambda x: float(str(x).replace(",","."))}
    s_d = {"sifra":str, "naziv":str, "jedinica":str, "cijena_eur": lambda x: float(str(x).replace(",",".")), "vrsta":str}

    mats=trake=fr=ftrake=usl=okov=oprema=dodatci=[]
    if url_mat:   mats   = from_csv_rows(pd.read_csv(url_mat).fillna("").to_dict(orient="records"), s_m)
    if url_trak:  trake  = from_csv_rows(pd.read_csv(url_trak).fillna("").to_dict(orient="records"), s_t)
    if url_fr:    fr     = from_csv_rows(pd.read_csv(url_fr).fillna("").to_dict(orient="records"), s_m)
    if url_ftrak: ftrake = from_csv_rows(pd.read_csv(url_ftrak).fillna("").to_dict(orient="records"), s_t)
    if url_usl:   usl    = from_csv_rows(pd.read_csv(url_usl).fillna("").to_dict(orient="records"), s_u)
    if url_okov:  okov   = from_csv_rows(pd.read_csv(url_okov).fillna("").to_dict(orient="records"), s_x)
    if url_oprema:oprema = from_csv_rows(pd.read_csv(url_oprema).fillna("").to_dict(orient="records"), s_x)
    if url_dodatci:dodatci=from_csv_rows(pd.read_csv(url_dodatci).fillna("").to_dict(orient="records"), s_d)

    return normalize_cjenik({
        "materijali": mats, "abs_trake": trake,
        "materijali_fronta": fr, "abs_trake_fronta": ftrake,
        "usluge": usl,
        "okov": okov, "oprema": oprema, "dodatci": dodatci
    })

CJE = None
if src == "Lokalni cjenik.json (default)":
    try:
        CJE = load_local(); st.sidebar.success("Učitano iz cjenik.json")
    except Exception as e:
        st.sidebar.error(f"Greška pri čitanju cjenik.json: {e}")
elif src == "Učitaj JSON (drag&drop)":
    up = st.sidebar.file_uploader("JSON s cjenikom", type=["json"])
    if up:
        try:
            CJE = load_from_uploaded(up.read())
            if st.sidebar.toggle("💾 Spremi kao cjenik.json", value=False):
                with open("cjenik.json","w",encoding="utf-8") as f:
                    json.dump(CJE, f, ensure_ascii=False, indent=2)
                st.sidebar.info("Spremljeno kao cjenik.json")
            st.sidebar.success("JSON učitan")
        except Exception as e:
            st.sidebar.error(f"Ne valja JSON: {e}")
    else:
        st.sidebar.info("Prevuci/odaberi JSON datoteku.")
elif src == "CSV URL-ovi (Google Sheets)":
    st.sidebar.caption("Očekivani stupci:")
    st.sidebar.code(
        "materijali/materijali_fronta: sifra, naziv, cijena_eur_po_m2\n"
        "abs_trake/abs_trake_fronta/usluge: sifra, naziv, cijena_eur_po_m\n"
        "okov/oprema: art_nr, naziv, dobavljac, jedinica, cijena_eur\n"
        "dodatci: sifra, naziv, jedinica, cijena_eur, vrsta (po kom|po m|po m2)"
    )
    url_mat   = st.sidebar.text_input("URL CSV – materijali (korpus)")
    url_trak  = st.sidebar.text_input("URL CSV – ABS trake (korpus)")
    url_fr    = st.sidebar.text_input("URL CSV – materijali fronta")
    url_ftrak = st.sidebar.text_input("URL CSV – ABS trake fronta")
    url_usl   = st.sidebar.text_input("URL CSV – usluge (rez/kant)")
    url_okov  = st.sidebar.text_input("URL CSV – OKOV")
    url_oprema= st.sidebar.text_input("URL CSV – OPREMA")
    url_dodatci = st.sidebar.text_input("URL CSV – DODATCI")
    if st.sidebar.button("🔗 Uvezi CSV", use_container_width=True):
        try:
            CJE = load_from_csv_urls(url_mat, url_trak, url_fr, url_ftrak, url_usl, url_okov, url_oprema, url_dodatci)
            st.sidebar.success("CSV uvezen")
        except Exception as e:
            st.sidebar.error(f"Greška pri čitanju CSV URL-ova: {e}")

if not CJE:
    st.stop()

# =============== Peek at pricebook ===============
with st.expander("📘 Pregled učitanog cjenika (klikni za detalje)"):
    c1, c2, c3 = st.columns(3)
    with c1: st.dataframe(CJE.get("materijali", []), use_container_width=True)
    with c2: st.dataframe(CJE.get("abs_trake", []), use_container_width=True)
    with c3: st.dataframe(CJE.get("usluge", []), use_container_width=True)
    c4, c5, c6 = st.columns(3)
    with c4: st.dataframe(CJE.get("okov", []), use_container_width=True)
    with c5: st.dataframe(CJE.get("oprema", []), use_container_width=True)
    with c6: st.dataframe(CJE.get("dodatci", []), use_container_width=True)

# =============== Peek at pricebook ===============
with st.expander("🧪 Dijagnostika cjenika (OKOV/OPREMA)"):
    raw_okov = CJE.get("okov", [])
    raw_oprema = CJE.get("oprema", [])
    miss_okov = [x for x in raw_okov if not str(x.get("art_nr") or "").strip()]
    miss_opr  = [x for x in raw_oprema if not str(x.get("art_nr") or "").strip()]
    st.write(f"OKOV učitano: {len(raw_okov)}  |  s valjanim 'art_nr': {len(raw_okov) - len(miss_okov)}")
    st.write(f"OPREMA učitano: {len(raw_oprema)}  |  s valjanim 'art_nr': {len(raw_oprema) - len(miss_opr)}")
    if miss_okov:
        st.warning("OKOV stavke bez 'art_nr' (ignorirane u padajućem izborniku):")
        st.dataframe(pd.DataFrame(miss_okov), use_container_width=True)
    if miss_opr:
        st.warning("OPREMA stavke bez 'art_nr' (ignorirane u padajućem izborniku):")
        st.dataframe(pd.DataFrame(miss_opr), use_container_width=True)

# =============== Wizard header ===============
st.markdown('<div class="sticky">🧮 <strong>Kalkulator Korpusa – Unified V5+</strong> &nbsp; <span class="badge">1) Dimenzije → 2) Materijali → 3) Fronta → 4) Okov/Oprema/Dodatci → 5) Rad & marža → 6) Sažetak</span></div>', unsafe_allow_html=True)

# =============== Helpers ===============
def fmt_eur(x): return f"{x:,.2f} €".replace(",", " ").replace(".", ",")
def fmt_m(x): return f"{x:,.2f} m".replace(",", " ").replace(".", ",")
def fmt_m2(x): return f"{x:,.3f} m²".replace(",", " ").replace(".", ",")
def mm2_to_m2(mm2: float) -> float: return mm2 / 1_000_000.0
def mm_to_m(mm: float) -> float: return mm / 1000.0

def kant_length_mm_longshort(w, d, long_cnt:int, short_cnt:int):
    long_e = max(w, d); short_e = min(w, d)
    long_cnt = max(0, min(2, int(long_cnt))); short_cnt = max(0, min(2, int(short_cnt)))
    return long_cnt * long_e + short_cnt * short_e

def kv_table(title, rows):
    st.markdown(f"#### {title}")
    html = ['<div class="section"><table class="kv">']
    for lab, val, *cls in rows:
        cls_attr = f' class="{cls[0]}"' if cls else ""
        html.append(f"<tr><th>{lab}</th><td{cls_attr}>{val}</td></tr>")
    html.append("</table></div>")
    st.markdown("\n".join(html), unsafe_allow_html=True)

# --- kratke oznake elemenata ---
def short_code_for(naziv: str) -> str:
    nz = (naziv or "").strip().lower()
    if nz.startswith("stranica"): return "Str"
    if nz.startswith("pod"): return "Pd"
    if nz.startswith("kapa_povez"): return "Pov"
    if nz.startswith("kapa"): return "Kp"
    if nz.startswith("polica"): return "Pol"
    if "leđa" in nz or "ledja" in nz: return "Ld"
    if nz.startswith("fronta"): return "Fr"
    if "haupt" in nz and "horizontalni" in nz: return "HptHor"
    if "haupt" in nz and "vertikalni" in nz: return "HptVer"
    return ""

# === helper za usklađen prikaz kantiranja (UI) s auto-pravilima ===
def auto_kant_counts(naziv: str, auto: bool, A: float, B: float, fallback_dugi: int, fallback_kratki: int):
    nz = (naziv or "").strip().lower()
    if not auto:
        return max(0, min(2, int(fallback_dugi or 0))), max(0, min(2, int(fallback_kratki or 0)))
    if nz.startswith("pod"):
        return (1, 2) if A >= B else (2, 1)
    if nz.startswith("kapa") and "povez" not in nz:
        return (1, 2) if A >= B else (2, 1)
    if "haupt" in nz and "horizontalni" in nz:
        return (1, 2) if A >= B else (2, 1)
    if nz.startswith("stranica") or ("haupt" in nz and "vertikalni" in nz):
        return 1, 1
    if nz.startswith("polica") or nz.startswith("fronta"):
        return 2, 2
    return max(0, min(2, int(fallback_dugi or 0))), max(0, min(2, int(fallback_kratki or 0)))

# ---- OKOV/OPREMA pick-list editor (select po Art. Nr. + Naziv + Dobavljač) ----
def picklist_editor(catalog: dict, keys: list, title: str, key: str):
    """
    Editor s padajućim izbornikom gdje se prikazuje i šifra (art_nr) i naziv (+ dobavljač).
    Korisnik bira npr. "OK-1001 — Pant 110° (Blum)", a mi to mapiramo natrag na art_nr.
    Robusno rukuje praznim redovima / NaN vrijednostima.
    """
    import math
    st.subheader(title)

    # Svježe opcije i mapa display -> art_nr
    options_disp = []
    DISP2ART = {}
    for art in sorted([k for k in catalog.keys() if k]):
        item = catalog.get(art) or {}
        naziv = str(item.get("naziv") or "").strip()
        dob = str(item.get("dobavljac") or "").strip()
        # Labela: ART — Naziv (Dobavljač)  [Dobavljač je opcionalan]
        label = f"{art} — {naziv}" + (f" ({dob})" if dob else "")
        options_disp.append(label)
        DISP2ART[label] = art

    if not options_disp:
        st.warning("Nema stavki u cjeniku za ovaj odjeljak (provjeri polje 'art_nr' u JSON-u).")

    # Start s jednim praznim retkom; korisnik može dodavati/brisati retke
    seed = [{"art_pick": "", "kolicina": 0}]
    edited = st.data_editor(
        pd.DataFrame(seed),
        num_rows="dynamic",
        hide_index=True,
        use_container_width=True,
        column_config={
            # U editoru prikazujemo kombinirani 'label' kao opciju
            "art_pick": st.column_config.SelectboxColumn("Art. Nr. / Naziv", options=[""] + options_disp),
            "kolicina": st.column_config.NumberColumn("Količina", min_value=0, step=1),
        },
        key=key,
    )

    def safe_str(x) -> str:
        if x is None: return ""
        if isinstance(x, float):
            try:
                if math.isnan(x): return ""
            except Exception:
                pass
        s = str(x).strip()
        return "" if s.lower() == "nan" else s

    def safe_int(x, default=0) -> int:
        if x is None: return default
        if isinstance(x, float):
            try:
                if math.isnan(x): return default
            except Exception:
                pass
        s = str(x).strip()
        if s == "" or s.lower() == "nan": return default
        try:
            return int(float(s))
        except Exception:
            return default

    out_rows = []
    preview_rows = []
    for _, r in edited.iterrows():
        pick = safe_str(r.get("art_pick"))
        qty = safe_int(r.get("kolicina"), 0)

        # preskoči prazan red ili količinu 0
        if not pick or qty <= 0:
            continue

        # mapiraj display natrag u art_nr
        art = DISP2ART.get(pick)
        if not art:
            # fallback: možda je korisnik direktno upisao art broj
            art = pick.split("—", 1)[0].strip()

        item = catalog.get(art, {})
        naziv = safe_str(item.get("naziv"))
        dob = safe_str(item.get("dobavljac"))
        jed = safe_str(item.get("jedinica")) or "kom"
        try:
            cij = float(item.get("cijena_eur") or 0.0)
        except Exception:
            cij = 0.0
        iznos = cij * qty

        out_rows.append({
            "kategorija": title.split()[0].upper().replace("🔩","").replace("🧰","").strip(),  # OKOV / OPREMA
            "art_nr": art, "naziv": naziv, "dobavljac": dob,
            "jedinica": jed, "cijena_eur": cij,
            "kolicina": qty, "iznos": iznos
        })
        preview_rows.append({
            "Art. Nr.": art, "Naziv": naziv, "Dobavljač": dob, "Jedinica": jed,
            "Cijena (€)": round(cij, 2), "Količina": qty, "Iznos (€)": round(iznos, 2)
        })

    if preview_rows:
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True)

    return out_rows



# =============== Dicts (pricebook) ===============
MATS   = {m["sifra"]: m for m in CJE.get("materijali", [])}
TRAK   = {t["sifra"]: t for t in CJE.get("abs_trake", [])}
FRONTS = {m["sifra"]: m for m in CJE.get("materijali_fronta", [])}
FTRAK  = {t["sifra"]: t for t in CJE.get("abs_trake_fronta", [])}
USLG   = {u["sifra"]: u for u in CJE.get("usluge", [])}

# OKOV/OPREMA: ključimo po Art. Nr.
OKOV   = {o["art_nr"]: o for o in CJE.get("okov", []) if o.get("art_nr")}
OPREMA = {o["art_nr"]: o for o in CJE.get("oprema", []) if o.get("art_nr")}
DODATCI= {d["sifra"]: d for d in CJE.get("dodatci", [])}

MATS_KEYS  = sorted(MATS.keys(),  key=lambda k: MATS[k].get("naziv",""))
TRAK_KEYS  = sorted(TRAK.keys(),  key=lambda k: TRAK[k].get("naziv",""))
FR_KEYS    = sorted(FRONTS.keys(),key=lambda k: FRONTS[k].get("naziv",""))
FTRAK_KEYS = sorted(FTRAK.keys(), key=lambda k: FTRAK[k].get("naziv",""))
OKOV_KEYS  = sorted(OKOV.keys())
OPREMA_KEYS= sorted(OPREMA.keys())
DOD_KEYS   = sorted(DODATCI.keys(),key=lambda k: DODATCI[k].get("naziv",""))

def extract_short(label: str) -> str:
    if not label: return ""
    tokens = [t for t in label.replace(",", " ").split() if any(ch.isalnum() for ch in t)]
    return " ".join(tokens[:2]) if tokens else label

ALL_MATS = {**MATS, **FRONTS}
ALL_TRAKS = {**TRAK, **FTRAK}

MAT_LABEL = {k: extract_short(v.get("naziv", k)) for k, v in ALL_MATS.items()}
MAT_BY_LABEL = {v: k for k, v in MAT_LABEL.items()}

TRAK_LABEL = {k: extract_short(v.get("naziv", k)) for k, v in ALL_TRAKS.items()}
TRAK_BY_LABEL = {v: k for k, v in TRAK_LABEL.items()}

# =============== Step 1: Dimenzije & sklapanje ===============
st.markdown("### 1) 📐 Osnovne dimenzije & sklapanje")
with st.container():
    c1, c2, c3, c4 = st.columns(4)
    with c1: W = st.number_input("Širina W (mm)", min_value=1, value=800, step=10)
    with c2: H = st.number_input("Visina H (mm)", min_value=1, value=720, step=10)
    with c3: D = st.number_input("Dubina D (mm)", min_value=1, value=320, step=10)
    with c4: t = st.number_input("Debljina ploče t (mm)", min_value=1, value=18, step=1)
    c5, c6, c7 = st.columns(3)
    with c5: include_back = st.checkbox("Leđa (HDF) uključena", value=True)
    with c6: pod_vrsta_vanjski = st.checkbox("Pod VANJSKI (preko stranica)", value=True)
    with c7: kapa_vrsta_vanjska = st.checkbox("Kapa VANJSKA (preko stranica)", value=False)
    c8, c9 = st.columns(2)
    with c8: n_police = st.number_input("Broj polica", min_value=0, value=2, step=1)
    with c9:
        include_kapa_povez = st.checkbox("Kapa_povez", value=False)
        kapa_povez_mode = st.radio("Širina Kapa_povez", ["Fiksno (mm)", "% dubine"], horizontal=True)
    d1, d2, d3 = st.columns(3)
    with d1: kapa_povez_sirina_mm = st.number_input("Kapa_povez – širina (mm)", min_value=1, value=150, step=1)
    with d2: kapa_povez_posto = st.slider("Kapa_povez – % dubine D", min_value=1, max_value=100, value=50)
    with d3:
        include_haupt_hor = st.checkbox("Haupt horizontalni", value=False)
        include_haupt_ver = st.checkbox("Haupt vertikalni", value=False)
        haupt_sirina_mm = st.number_input("Širina haupta (mm)", min_value=1, value=80, step=1)

# =============== Step 2: Materijali & usluge ===============
st.markdown("### 2) 🧱 Materijali & usluge")
with st.container():
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        default_mat = st.selectbox("Materijal korpusa", MATS_KEYS,
            format_func=lambda k: f'{k} – {MATS[k]["naziv"]} ({MATS[k]["cijena_eur_po_m2"]:.2f} €/m²)')
    with m2:
        default_traka = st.selectbox("ABS traka korpusa", TRAK_KEYS,
            format_func=lambda k: f'{k} – {TRAK[k]["naziv"]} ({TRAK[k]["cijena_eur_po_m"]:.2f} €/m)')
    with m3:
        default_mat_fr = st.selectbox("Materijal fronte", FR_KEYS,
            format_func=lambda k: f'{k} – {FRONTS[k]["naziv"]} ({FRONTS[k]["cijena_eur_po_m2"]:.2f} €/m²)')
    with m4:
        default_traka_fr = st.selectbox("ABS traka fronte", FTRAK_KEYS,
            format_func=lambda k: f'{k} – {FTRAK[k]["naziv"]} ({FTRAK[k]["cijena_eur_po_m"]:.2f} €/m)')
    u1, u2 = st.columns(2)
    with u1:
        rez_keys = [k for k,v in USLG.items() if "cijena_eur_po_m" in v]
        rez_usl = st.selectbox("Usluga rezanja (€/m)", sorted(rez_keys),
            format_func=lambda k: f'{k} – {USLG[k]["naziv"]} ({USLG[k]["cijena_eur_po_m"]:.2f} €/m)')
    with u2:
        kant_keys = [k for k,v in USLG.items() if "cijena_eur_po_m" in v]
        kant_usl = st.selectbox("Usluga kantiranja (€/m)", sorted(kant_keys),
            format_func=lambda k: f'{k} – {USLG[k]["naziv"]} ({USLG[k]["cijena_eur_po_m"]:.2f} €/m)')

# =============== Step 3: Fronta ===============
st.markdown("### 3) 🚪 Fronta")
with st.container():
    include_fronta = st.checkbox("Dodaj frontu", value=False)
    f1, f2 = st.columns(2)
    with f1: fronta_tip = st.selectbox("Tip fronte", ["Jednokrilna", "Dvokrilna"])
    with f2: fronta_montaza = st.selectbox("Montaža", ["Unutarnja (u korpusu)", "Vanjska (preko korpusa)"])
    g1, g2, g3 = st.columns(3)
    with g1: razmak_hor = st.number_input("Razmak horiz. (mm)", min_value=0.0, value=2.0, step=0.5)
    with g2: razmak_ver = st.number_input("Razmak vert. (mm)", min_value=0.0, value=2.0, step=0.5)
    with g3: razmak_srednji = st.number_input("Srednji razmak (dvokrilna) (mm)", min_value=0.0, value=2.0, step=0.5)
    h1, h2 = st.columns(2)
    with h1: preklop_hor = st.number_input("Preklop horiz. (mm)", min_value=0.0, value=0.0, step=0.5)
    with h2: preklop_ver = st.number_input("Preklop vert. (mm)", min_value=0.0, value=0.0, step=0.5)

# =============== Step 4: OKOV / OPREMA / DODATCI ===============
st.markdown("### 4) 🔩 OKOV • 🧰 OPREMA • 🧱 Dodatci (ručni unos dimenzija)")

# --- OKOV s padajućim izbornikom (po Art. Nr.) ---
okov_rows = picklist_editor(OKOV, OKOV_KEYS, "🔩 OKOV", key="okov_editor")

# --- OPREMA s padajućim izbornikom (po Art. Nr.) ---
oprema_rows = picklist_editor(OPREMA, OPREMA_KEYS, "🧰 OPREMA", key="oprema_editor")

# --- DODATCI (ručni unos dimenzija) – po kom / po m / po m2 ---
st.subheader("🧱 Dodatci (ručni unos dimenzija)")
dodatci_template = []
for k in DOD_KEYS:
    row = {**DODATCI[k], "sifra":k}
    row.setdefault("vrsta","po kom")   # po kom | po m | po m2
    row["A_mm"] = 0
    row["B_mm"] = 0
    row["kom"] = 0
    dodatci_template.append(row)
if not dodatci_template:
    dodatci_template = [{"sifra":"DD-001","naziv":"Dodatni element – placeholder","jedinica":"po kom","cijena_eur":10.0,"vrsta":"po kom","A_mm":0,"B_mm":0,"kom":0}]

edited_dodatci = st.data_editor(
    pd.DataFrame(dodatci_template)[["sifra","naziv","vrsta","jedinica","cijena_eur","A_mm","B_mm","kom"]],
    hide_index=True, use_container_width=True,
    column_config={
        "sifra": st.column_config.TextColumn("Šifra", disabled=True),
        "naziv": st.column_config.TextColumn("Naziv", disabled=True),
        "vrsta": st.column_config.SelectboxColumn("Vrsta obračuna", options=["po kom","po m","po m2"]),
        "jedinica": st.column_config.TextColumn("Jedinica", disabled=True),
        "cijena_eur": st.column_config.NumberColumn("Cijena (€)", format="%.2f"),
        "A_mm": st.column_config.NumberColumn("Dim A (mm)", min_value=0, step=1),
        "B_mm": st.column_config.NumberColumn("Dim B (mm)", min_value=0, step=1),
        "kom": st.column_config.NumberColumn("Kom", min_value=0, step=1),
    }
)

dodatci_rows = []
for _, r in edited_dodatci.iterrows():
    kom = int(r["kom"]) if r["kom"] else 0
    if kom <= 0:
        continue
    A = float(r["A_mm"] or 0); B = float(r["B_mm"] or 0)
    vrsta = str(r["vrsta"] or "po kom").strip().lower()
    jedinica = r["jedinica"] or ("kom" if vrsta=="po kom" else ("m" if vrsta=="po m" else "m²"))
    cij = float(r["cijena_eur"] or 0.0)

    if vrsta == "po m2":
        kolicina_obracun = mm2_to_m2(A*B) * kom
    elif vrsta == "po m":
        kolicina_obracun = mm_to_m(max(A,B)+min(A,B)) * kom
    else:  # po kom
        kolicina_obracun = kom

    iznos = cij * kolicina_obracun
    dodatci_rows.append({
        "kategorija":"DODATAK", "sifra": r["sifra"], "naziv": r["naziv"],
        "vrsta": vrsta, "jedinica": jedinica,
        "A_mm": int(A), "B_mm": int(B), "kom": kom,
        "obračun_količina": round(kolicina_obracun, 3),
        "cijena_eur": cij, "iznos": iznos
    })

# =============== Step 5: Rad i marža ===============
st.markdown("### 5) 🛠️ Rad i marža")
with st.container():
    r1, r2, r3 = st.columns(3)
    with r1:
        h_tp = st.number_input("Tehnička priprema – sati", min_value=0.0, value=0.5, step=0.25)
        r_tp = st.number_input("Cijena rada TP (€/h)", min_value=0.0, value=28.0, step=1.0)
    with r2:
        h_cnc = st.number_input("CNC i strojna obrada – sati", min_value=0.0, value=0.8, step=0.25)
        r_cnc = st.number_input("Cijena rada CNC (€/h)", min_value=0.0, value=35.0, step=1.0)
    with r3:
        h_skl = st.number_input("Sklapanje & montaža – sati", min_value=0.0, value=0.7, step=0.25)
        r_skl = st.number_input("Cijena rada SKL (€/h)", min_value=0.0, value=30.0, step=1.0)
    rp1, rp2, rp3 = st.columns(3)
    with rp1: h_pak = st.number_input("Pakiranje – sati", min_value=0.0, value=0.3, step=0.25)
    with rp2: r_pak = st.number_input("Cijena rada PAK (€/h)", min_value=0.0, value=22.0, step=1.0)
    with rp3:
        use_waste = st.checkbox("Uključi otpad (%)", value=True)
        waste_pct = st.number_input("Postotak otpada (%)", min_value=0.0, value=8.0, step=0.5)
    o1, o2 = st.columns(2)
    with o1: use_markup = st.checkbox("Uključi maržu (%)", value=False)
    with o2: markup_pct = st.number_input("Postotak marže (%)", min_value=0.0, value=15.0, step=0.5)
    rok_dani = st.number_input("Planirana isporuka (dana od narudžbe)", min_value=0, value=30, step=1)

# =============== Calculation functions ===============
def derive_rows(W,H,D,t,n_police, include_back, default_mat, default_traka, pod_vrsta_vanjski, kapa_vrsta_vanjska,
                include_kapa_povez, kapa_povez_mode, kapa_povez_sirina_mm, kapa_povez_posto,
                include_fronta, fronta_tip, fronta_montaza, razmak_hor, razmak_ver, razmak_srednji,
                preklop_hor, preklop_ver, default_mat_fr, default_traka_fr,
                include_haupt_hor, include_haupt_ver, haupt_sirina_mm):
    inner_w = max(W - 2*t, 0)
    side_h = max(H - (t if pod_vrsta_vanjski else 0) - (t if kapa_vrsta_vanjska else 0), 1)

    rows = []
    rows.append({"naziv":"Stranica", "mat": default_mat, "traka": default_traka,
                 "A_mm": side_h, "B_mm": D, "kom":2, "kant_dugi":1, "kant_kratki":1, "auto": True})

    # Ako koristimo Kapa_povez, preskačemo klasičnu "Kapu"
    if not include_kapa_povez:
        kapa_w = W if kapa_vrsta_vanjska else inner_w
        rows.append({"naziv":"Kapa", "mat": default_mat, "traka": default_traka,
                     "A_mm": kapa_w, "B_mm": D, "kom":1, "kant_dugi":1, "kant_kratki":2, "auto": True})  # pravilo kao Pod

    pod_w = W if pod_vrsta_vanjski else inner_w
    rows.append({"naziv":"Pod", "mat": default_mat, "traka": default_traka,
                 "A_mm": pod_w, "B_mm": D, "kom":1, "kant_dugi":1, "kant_kratki":2, "auto": True})

    if n_police > 0:
        pol_w = max(inner_w - 2, 1); pol_d = max(D - 10, 1)
        rows.append({"naziv":"Polica", "mat": default_mat, "traka": default_traka,
                     "A_mm": pol_w, "B_mm": pol_d, "kom": int(n_police),
                     "kant_dugi":2, "kant_kratki":2, "auto": True})

    if include_kapa_povez:
        width = int(round(D * (kapa_povez_posto / 100.0))) if kapa_povez_mode == "% dubine" else int(kapa_povez_sirina_mm)
        width = max(1, min(width, int(D)))
        rows.append({"naziv":"Kapa_povez", "mat": default_mat, "traka": default_traka,
                     "A_mm": inner_w, "B_mm": width, "kom":2, "kant_dugi":2, "kant_kratki":0, "auto": False})

    if include_back:
        rows.append({"naziv":"Leđa (HDF)", "mat": "HDF-001", "traka": default_traka,
                     "A_mm": max(W-2,1), "B_mm": max(H-2,1), "kom":1,
                     "kant_dugi":0, "kant_kratki":0, "auto": False})

    # FRONT
    if include_fronta:
        if fronta_montaza.startswith("Unutarnja"):
            target_w = max(inner_w - razmak_hor, 1); target_h = max(H - razmak_ver, 1)
            ukupna_sirina = max(inner_w - razmak_hor, 1)
        else:
            target_w = W + preklop_hor; target_h = H + preklop_ver
            ukupna_sirina = W + preklop_hor
        if fronta_tip == "Jednokrilna":
            rows.append({"naziv":"Fronta", "mat": default_mat_fr, "traka": default_traka_fr,
                         "A_mm": target_h, "B_mm": target_w, "kom":1,
                         "kant_dugi":2, "kant_kratki":2, "auto": True})
        else:
            left_w = max((ukupna_sirina - razmak_srednji)/2.0, 1)
            for side in ("L","D"):
                rows.append({"naziv": f"Fronta {side}", "mat": default_mat_fr, "traka": default_traka_fr,
                             "A_mm": target_h, "B_mm": int(round(left_w)), "kom":1,
                             "kant_dugi":2, "kant_kratki":2, "auto": True})

    # HAUPT — dimenzije i dodavanje (B = D - 10)
    haupt_depth = max(D - 10, 1)

    if include_haupt_hor:
        rows.append({"naziv":"Haupt Horizontalni", "mat": default_mat, "traka": default_traka,
                     "A_mm": max(inner_w,1), "B_mm": haupt_depth,
                     "kom":1, "kant_dugi":1, "kant_kratki":2, "auto": True})

    if include_haupt_ver:
        hpt_ver_len = max(side_h, 1)  # unutarnja visina
        rows.append({"naziv":"Haupt Vertikalni", "mat": default_mat, "traka": default_traka,
                     "A_mm": int(hpt_ver_len), "B_mm": haupt_depth,
                     "kom":1, "kant_dugi":1, "kant_kratki":1, "auto": True})

    return rows

def calculate(report_rows, rez_usl, kant_usl, MATS, TRAK, FRONTS, FTRAK, USLG):
    rez_cij_m = USLG[rez_usl]["cijena_eur_po_m"]
    kant_usl_cij_m = USLG[kant_usl]["cijena_eur_po_m"]

    total_area_m2 = total_rezanje_m = total_kant_m = 0.0
    cijena_mat_eur = cijena_kant_traka_eur = cijena_kant_usl_eur = cijena_rez_eur = 0.0
    iveral_area_m2 = iveral_eur = 0.0
    hdf_area_m2 = hdf_eur = 0.0

    def rezanje_rule(A, B):
        return mm_to_m(max(A, B) + min(A, B))

    report = []
    for r in report_rows:
        A = float(r["A_mm"]); B = float(r["B_mm"]); k = int(r["kom"])
        rez_m_tot = rezanje_rule(A, B) * k
        naziv = str(r["naziv"]).lower()
        auto = bool(r.get("auto", False))

        # --- KANTIRANJE (auto pravila) ---
        if auto and naziv.startswith("pod"):
            kant_mm_kom = (A * 1 + B * 2) if A >= B else (A * 2 + B * 1)
        elif auto and naziv.startswith("kapa") and "povez" not in naziv:
            kant_mm_kom = (A * 1 + B * 2) if A >= B else (A * 2 + B * 1)
        elif auto and ("haupt" in naziv and "horizontalni" in naziv):
            kant_mm_kom = (A * 1 + B * 2) if A >= B else (A * 2 + B * 1)
        elif auto and (naziv.startswith("stranica") or ("haupt" in naziv and "vertikalni" in naziv)):
            kant_mm_kom = max(A, B) + min(A, B)
        elif auto and (naziv.startswith("polica") or naziv.startswith("fronta")):
            kant_mm_kom = 2 * max(A, B) + 2 * min(A, B)
        else:
            kant_mm_kom = kant_length_mm_longshort(A, B, int(r.get("kant_dugi", 0)), int(r.get("kant_kratki", 0)))

        kant_m_tot = mm_to_m(kant_mm_kom) * k
        area_m2_tot = mm2_to_m2(A * B) * k

        # --- LABELI: kratke šifre iz naziva ---
        mat_obj = MATS.get(r["mat"]) or FRONTS.get(r["mat"]) or {}
        traka_obj = TRAK.get(r["traka"]) or FTRAK.get(r["traka"]) or {}

        mat_label = extract_short(mat_obj.get("naziv", str(r["mat"])))
        traka_label = extract_short(traka_obj.get("naziv", str(r["traka"])))

        mat_price = (mat_obj.get("cijena_eur_po_m2") or 0.0)
        traka_price = (traka_obj.get("cijena_eur_po_m") or 0.0)

        mat_cij = mat_price * area_m2_tot
        traka_cij = traka_price * kant_m_tot
        rez_cij = rez_cij_m * rez_m_tot
        kant_usl_e = kant_usl_cij_m * kant_m_tot

        if r["mat"] == "HDF-001":
            hdf_area_m2 += area_m2_tot; hdf_eur += mat_cij
        else:
            iveral_area_m2 += area_m2_tot; iveral_eur += mat_cij

        total_area_m2 += area_m2_tot
        total_rezanje_m += rez_m_tot
        total_kant_m += kant_m_tot
        cijena_mat_eur += mat_cij
        cijena_kant_traka_eur += traka_cij
        cijena_kant_usl_eur += kant_usl_e
        cijena_rez_eur += rez_cij

        report.append({
            "Naziv": r["naziv"],
            "Oznaka": short_code_for(r["naziv"]),
            "Mat": mat_label,
            "Traka": traka_label,
            "A (mm)": int(A),
            "B (mm)": int(B),
            "Kom": k,
            "Kant m": round(kant_m_tot, 3),
            "Rezanje m": round(rez_m_tot, 3),
            "Površina m²": round(area_m2_tot, 3),
            "€ Materijal": round(mat_cij, 2),
            "€ Traka": round(traka_cij, 2),
            "€ Usl. kant": round(kant_usl_e, 2),
            "€ Rezanje": round(rez_cij, 2),
            "€ Element (ukupno)": round(mat_cij + traka_cij + rez_cij + kant_usl_e, 2),
        })

    metrics = dict(
        total_area_m2=total_area_m2,
        total_rezanje_m=total_rezanje_m,
        total_kant_m=total_kant_m,
        cijena_mat_eur=cijena_mat_eur,
        cijena_kant_traka_eur=cijena_kant_traka_eur,
        cijena_kant_usl_eur=cijena_kant_usl_eur,
        cijena_rez_eur=cijena_rez_eur,
        iveral_area_m2=iveral_area_m2,
        iveral_eur=iveral_eur,
        hdf_area_m2=hdf_area_m2,
        hdf_eur=hdf_eur,
    )
    return report, metrics

def materials_services_summary(metrics, use_waste, waste_pct):
    eur_mats_total = metrics['cijena_mat_eur']; eur_trake = metrics['cijena_kant_traka_eur']
    eur_rezanje = metrics['cijena_rez_eur']; eur_kant_usl = metrics['cijena_kant_usl_eur']
    eur_waste = (waste_pct/100.0)*(eur_mats_total + eur_trake) if use_waste else 0.0
    subtotal = eur_mats_total + eur_trake + eur_rezanje + eur_kant_usl + eur_waste
    kv_table("📊 Materijal + usluge (korpus)", [
        ("m² iveral", fmt_m2(metrics['iveral_area_m2'])),
        ("€ iveral", fmt_eur(metrics['iveral_eur'])),
        ("m² HDF", fmt_m2(metrics['hdf_area_m2'])),
        ("€ HDF", fmt_eur(metrics['hdf_eur'])),
        ("m² ukupno", fmt_m2(metrics['total_area_m2'])),
        ("€ materijal ukupno", fmt_eur(eur_mats_total)),
        ("Rezanje (m)", fmt_m(metrics['total_rezanje_m'])),
        ("€ rezanje", fmt_eur(eur_rezanje)),
        ("Kantiranje (m)", fmt_m(metrics['total_kant_m'])),
        ("€ trake", fmt_eur(eur_trake)),
        ("€ usluga kantiranja", fmt_eur(eur_kant_usl)),
        ("€ otpad", fmt_eur(eur_waste)),
        ("Materijal + usluge + otpad", fmt_eur(subtotal), "total"),
    ])
    return subtotal, eur_waste

def extras_totals(okov_rows, oprema_rows, dodatci_rows):
    sum_okov = sum(r["iznos"] for r in okov_rows)
    sum_oprema = sum(r["iznos"] for r in oprema_rows)
    sum_dodatci = sum(r["iznos"] for r in dodatci_rows)
    total = sum_okov + sum_oprema + sum_dodatci
    kv_table("🔩🧰🧱 Okov + Oprema + Dodatci", [
        ("OKOV", fmt_eur(sum_okov)),
        ("OPREMA", fmt_eur(sum_oprema)),
        ("Dodatci (ručni)", fmt_eur(sum_dodatci)),
        ("UKUPNO (okov/oprema/dodatci)", fmt_eur(total), "total"),
    ])
    return total

def labor_total_calc(h_tp,r_tp,h_cnc,r_cnc,h_skl,r_skl,h_pak,r_pak):
    return h_tp*r_tp + h_cnc*r_cnc + h_skl*r_skl + h_pak*r_pak

def final_summary_grand(mats_services_total, extras_total, labor_total, use_markup, markup_pct):
    pre_markup = mats_services_total + extras_total + labor_total
    eur_markup = (markup_pct/100.0)*pre_markup if use_markup else 0.0
    ukupno = pre_markup + eur_markup
    kv_table("🧾 Završni zbir", [
        ("Materijal + usluge + otpad (korpus)", fmt_eur(mats_services_total)),
        ("Okov + Oprema + Dodatci", fmt_eur(extras_total)),
        ("Rad (sati × €/h)", fmt_eur(labor_total)),
        ("Zbroj (prije marže)", fmt_eur(pre_markup)),
        ("Marža", fmt_eur(eur_markup) + (f"  ({markup_pct:.1f} %)" if use_markup else "  (0 %)")),
        ("UKUPNO", fmt_eur(ukupno), "total"),
    ])
    return ukupno

# -------- PDF Export (s NAZIV naslovom) --------
def _pdf_title_korpus(W, H, D, include_fronta, fronta_tip, fronta_montaza):
    dims = f"Korpus H={int(H)}mm × W={int(W)}mm × D={int(D)}mm"
    if include_fronta:
        krila = "2F" if str(fronta_tip).lower().startswith("dvokrilna") else "1F"
        mont = "Unutarnja" if str(fronta_montaza).lower().startswith("unut") else "Vanjska"
        return f"{dims} — {krila} ({mont})"
    return f"{dims} — bez fronte"

def build_full_pdf(report, metrics, mats_services_total, extras_total, labor_total, use_markup, markup_pct,
                   W, H, D, n_police, waste_pct, rok_dani,
                   include_back, pod_vrsta_vanjski, kapa_vrsta_vanjska, include_kapa_povez,
                   include_fronta, fronta_tip, fronta_montaza):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    import io, datetime

    title = _pdf_title_korpus(W, H, D, include_fronta, fronta_tip, fronta_montaza)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="H1", fontSize=16, leading=20, spaceAfter=6))
    styles.add(ParagraphStyle(name="H2", fontSize=12, leading=16, spaceBefore=8, spaceAfter=4))

    elems = []
    elems.append(Paragraph(title, styles["H1"]))
    elems.append(Paragraph(datetime.datetime.now().strftime("%d.%m.%Y."), styles["Normal"]))
    elems.append(Spacer(1, 6))

    kv = [
        ["Stavka", "Vrijednost"],
        ["Širina (W)", f"{int(W)} mm"],
        ["Visina (H)", f"{int(H)} mm"],
        ["Dubina (D)", f"{int(D)} mm"],
        ["Broj polica", f"{int(n_police)}"],
        ["Leđa HDF", "DA" if include_back else "NE"],
        ["Pod VANJSKI", "DA" if pod_vrsta_vanjski else "NE"],
        ["Kapa VANJSKA", "DA" if kapa_vrsta_vanjska else "NE"],
        ["Kapa_povez", "DA" if include_kapa_povez else "NE"],
        ["Planirana isporuka", f"{int(rok_dani)} dana"],
    ]
    t_kv = Table(kv, colWidths=[70 * mm, 80 * mm])
    t_kv.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
    ]))
    elems += [Paragraph("📐 Osnovne postavke", styles["H2"]), t_kv, Spacer(1, 6)]

    if report:
        header = ["Naziv","Oznaka","Mat","Traka","A (mm)","B (mm)","Kom",
                  "Kant m","Rezanje m","Površina m²","€ Materijal","€ Traka","€ Usl. kant","€ Rezanje","€ Element (ukupno)"]
        data = [header] + [[str(r.get(k, "")) for k in header] for r in report]
        cw = [28*mm,12*mm,18*mm,18*mm,14*mm,14*mm,10*mm,16*mm,16*mm,16*mm,18*mm,16*mm,18*mm,16*mm,22*mm]
        t_rep = Table(data, repeatRows=1, colWidths=cw)
        t_rep.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#e5e7eb")),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f3f4f6")),
            ("ALIGN", (1,1), (-1,-1), "CENTER"),
            ("ALIGN", (0,1), (0,-1), "LEFT"),
        ]))
        elems += [Paragraph("🧾 Elementi i troškovi (korpus)", styles["H2"]), t_rep, Spacer(1, 6)]

    ms = metrics
    eur_waste = (waste_pct/100.0)*(ms['cijena_mat_eur'] + ms['cijena_kant_traka_eur'])
    mats_rows = [
        ["m² iveral", f"{ms['iveral_area_m2']:.3f}", f"{ms['iveral_eur']:.2f}"],
        ["m² HDF", f"{ms['hdf_area_m2']:.3f}", f"{ms['hdf_eur']:.2f}"],
        ["m² ukupno", f"{ms['total_area_m2']:.3f}", f"{ms['cijena_mat_eur']:.2f}"],
        ["Rezanje (m)", f"{ms['total_rezanje_m']:.3f}", f"{ms['cijena_rez_eur']:.2f}"],
        ["Kantiranje (m)", f"{ms['total_kant_m']:.3f}", f"{ms['cijena_kant_traka_eur']:.2f}"],
        ["€ usluga kantiranja", "", f"{ms['cijena_kant_usl_eur']:.2f}"],
        ["€ otpad", "", f"{eur_waste:.2f}"],
        ["Materijal + usluge + otpad", "", f"{mats_services_total:.2f}"],
    ]
    t_ms = Table([["Stavka","Količina","Iznos (€)"]] + mats_rows, colWidths=[80*mm, 35*mm, 45*mm])
    t_ms.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#e5e7eb")),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#f3f4f6")),
        ("ALIGN", (1,1), (-1,-1), "CENTER"),
        ("ALIGN", (0,0), (0,-1), "LEFT"),
    ]))
    elems += [Paragraph("📊 Materijal + usluge (korpus)", styles["H2"]), t_ms, Spacer(1, 6)]

    pre_markup = mats_services_total + extras_total + labor_total
    eur_markup = pre_markup * (markup_pct/100.0) if use_markup else 0.0
    ukupno_pdf = pre_markup + eur_markup
    t_fin = Table([
        ["Okov + Oprema + Dodatci", f"{extras_total:.2f}"],
        ["Rad (sati × €/h)", f"{labor_total:.2f}"],
        ["Zbroj prije marže", f"{pre_markup:.2f}"],
        [f"Marža ({markup_pct:.1f}% )" if use_markup else "Marža (0%)", f"{eur_markup:.2f}"],
        ["UKUPNO", f"{ukupno_pdf:.2f}"],
    ], colWidths=[100*mm, 60*mm])
    t_fin.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#e5e7eb")),
        ("BACKGROUND", (0,4), (-1,4), colors.HexColor("#effbf1")),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("ALIGN", (0,0), (0,-1), "LEFT"),
    ]))
    elems += [Paragraph("🧾 Završni zbir", styles["H2"]), t_fin]

    doc.build(elems)
    return buf.getvalue()

# =============== EXCEL EXPORT (TOP-LEVEL FUNKCIJA) ===============
def build_xlsx_kantiranje(report_rows, source_rows, okov_rows, oprema_rows, dodatci_rows):
    import io
    import pandas as pd
    import importlib.util

    if not report_rows:
        return None, "Nema podataka za izvoz (report je prazan)."

    # Helper: (dugi_cnt, kratki_cnt) po retku na temelju auto pravila ili ručnih postavki
    def kant_counts_for_row(src_row):
        naziv = str(src_row.get("naziv", "")).lower()
        auto = bool(src_row.get("auto", False))

        if naziv.startswith("kapa_povez"):
            return 2, 0

        def ab(row):
            try:
                A = float(row.get("A_mm", 0)); B = float(row.get("B_mm", 0))
            except Exception:
                A = 0; B = 0
            return A, B

        if naziv.startswith("pod"):
            A,B = ab(src_row); return (1, 2) if A >= B else (2, 1)
        if naziv.startswith("kapa") and "povez" not in naziv:
            A,B = ab(src_row); return (1, 2) if A >= B else (2, 1)
        if "haupt" in naziv and "horizontalni" in naziv:
            A,B = ab(src_row); return (1, 2) if A >= B else (2, 1)
        if auto and (naziv.startswith("stranica") or ("haupt" in naziv and "vertikalni" in naziv)):
            return 1, 1
        if auto and (naziv.startswith("polica") or naziv.startswith("fronta")):
            return 2, 2

        d = int(src_row.get("kant_dugi", 0) or 0)
        k = int(src_row.get("kant_kratki", 0) or 0)
        return max(0, min(2, d)), max(0, min(2, k))

    # Kategorizacija površine
    def classify_surface(src_row):
        mat = str(src_row.get("mat", "")).upper()
        naziv = str(src_row.get("naziv", "")).lower()
        if mat == "HDF-001":
            return "Leđa HDF"
        if naziv.startswith("fronta"):
            return "Fronte Iveral"
        return "Korpusi Iveral"

    # Core retci
    rows_core = []
    korp_iveral_m2 = fronte_iveral_m2 = hdf_m2 = total_rez_m = 0.0

    for i, rep in enumerate(report_rows):
        src = source_rows[i] if i < len(source_rows) else {}
        d_cnt, k_cnt = kant_counts_for_row(src)
        oznaka_k = f"{k_cnt}K {d_cnt}D"
        povrsina_m2 = float(rep.get("Površina m²", 0) or 0)
        rezanje_m = float(rep.get("Rezanje m", 0) or 0)

        cat = classify_surface(src)
        if cat == "Korpusi Iveral":
            korp_iveral_m2 += povrsina_m2
        elif cat == "Fronte Iveral":
            fronte_iveral_m2 += povrsina_m2
        else:
            hdf_m2 += povrsina_m2
        total_rez_m += rezanje_m

        rows_core.append({
            "Naziv": rep.get("Naziv", ""),
            "Oznaka": rep.get("Oznaka", ""),
            "Mat": rep.get("Mat", ""),
            "Traka": rep.get("Traka", ""),
            "A (mm)": rep.get("A (mm)", ""),
            "B (mm)": rep.get("B (mm)", ""),
            "Kom": rep.get("Kom", ""),
            "Kratke strane (K)": k_cnt,
            "Duge strane (D)": d_cnt,
            "Oznaka kantiranja": oznaka_k,
            "Kant m": rep.get("Kant m", 0),
            "Površina m²": povrsina_m2,
            "Rezanje m": rezanje_m,
        })

    order_mats, seen = [], set()
    for r in rows_core:
        m = r.get("Mat", "")
        if m not in seen:
            seen.add(m)
            order_mats.append(m)

    df_core = pd.DataFrame(rows_core)
    gb = df_core.groupby("Mat", dropna=False).agg({
        "Kant m": "sum",
        "Površina m²": "sum",
        "Rezanje m": "sum"
    }).reset_index()
    per_mat_totals = {row["Mat"]: {
        "Kant m": float(row["Kant m"]),
        "Površina m²": float(row["Površina m²"]),
        "Rezanje m": float(row["Rezanje m"]),
    } for _, row in gb.iterrows()}

    cols_order = ["Naziv","Oznaka","Mat","Traka","A (mm)","B (mm)","Kom",
                  "Kratke strane (K)","Duge strane (D)","Oznaka kantiranja",
                  "Kant m","Površina m²","Rezanje m"]

    rows_display = []
    subtotal_row_indices = []
    for mat in order_mats:
        for r in rows_core:
            if r.get("Mat", "") == mat:
                rows_display.append({c: r.get(c, "") for c in cols_order})
        subt = per_mat_totals.get(mat, {"Kant m":0, "Površina m²":0, "Rezanje m":0})
        subtotal_row = {c: "" for c in cols_order}
        subtotal_row["Naziv"] = f"UKUPNO – {mat}"
        subtotal_row["Mat"] = mat
        subtotal_row["Kant m"] = subt["Kant m"]
        subtotal_row["Površina m²"] = subt["Površina m²"]
        subtotal_row["Rezanje m"] = subt["Rezanje m"]
        rows_display.append(subtotal_row)
        subtotal_row_indices.append(len(rows_display))

    df = pd.DataFrame(rows_display, columns=cols_order)

    sum_by_traka = (
        df_core[["Traka","Kant m"]]
        .groupby("Traka", dropna=False)["Kant m"].sum()
        .reset_index()
        .rename(columns={"Kant m":"Kant m ukupno"})
    )

    total_materials_m2 = korp_iveral_m2 + fronte_iveral_m2 + hdf_m2

    # Engine
    import importlib.util
    has_openpyxl   = importlib.util.find_spec("openpyxl")  is not None
    has_xlsxwriter = importlib.util.find_spec("xlsxwriter") is not None
    engine = "openpyxl" if has_openpyxl else ("xlsxwriter" if has_xlsxwriter else None)
    if engine is None:
        return None, "Nedostaje engine za Excel. Instaliraj: pip install openpyxl (ili xlsxwriter)."

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine=engine) as writer:
        # ========== Sheet 1: Elementi_kantiranje ==========
        df.to_excel(writer, index=False, sheet_name="Elementi_kantiranje")
        ws1 = writer.sheets["Elementi_kantiranje"]

        grand_total_kant = float(df_core["Kant m"].sum())
        grand_total_m2   = float(df_core["Površina m²"].sum())
        grand_total_rez  = float(df_core["Rezanje m"].sum())

        if engine == "xlsxwriter":
            bold = writer.book.add_format({"bold": True})
            grey = writer.book.add_format({"bold": True, "bg_color": "#f3f4f6"})
            green = writer.book.add_format({"bold": True, "bg_color": "#eef7ee"})
            left = writer.book.add_format({"align":"left"})
            center = writer.book.add_format({"align":"center"})
            num3 = writer.book.add_format({"num_format":"0.000", "align":"center"})
            widths = {"Naziv": 28, "Oznaka":10, "Mat":18, "Traka":18, "A (mm)":12, "B (mm)":12, "Kom":8,
                      "Kratke strane (K)":16, "Duge strane (D)":16, "Oznaka kantiranja":16,
                      "Kant m":14, "Površina m²":14, "Rezanje m":14}
            for i, name in enumerate(df.columns):
                ws1.set_column(i, i, widths.get(name, 14))
            ws1.autofilter(0, 0, len(df), len(df.columns)-1)
            ws1.freeze_panes(1, 1)
            ws1.set_column(0, 0, widths["Naziv"], left)
            ws1.set_column(1, len(df.columns)-1, None, center)
            col_idx = {n:i for i,n in enumerate(df.columns)}
            for col in ["Kant m","Rezanje m","Površina m²"]:
                ws1.set_column(col_idx[col], col_idx[col], widths[col], num3)
            total_row_excel = len(df) + 1
            ws1.write(total_row_excel, 0, "UKUPNO – SVI MATERIJALI", bold)
            ws1.write_number(total_row_excel, col_idx["Kant m"], grand_total_kant, green)
            ws1.write_number(total_row_excel, col_idx["Površina m²"], grand_total_m2, green)
            ws1.write_number(total_row_excel, col_idx["Rezanje m"], grand_total_rez, green)
            for idx in subtotal_row_indices:
                excel_row = idx
                ws1.set_row(excel_row, 12, grey)
        else:
            from openpyxl.styles import Alignment, Font, PatternFill
            ws1.freeze_panes = "B2"
            ws1.auto_filter.ref = ws1.dimensions
            for cell in ws1[1]:
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
                if row and row[0].value is not None:
                    row[0].alignment = Alignment(horizontal="left")
                for c in row[1:]:
                    c.alignment = Alignment(horizontal="center")
            last = ws1.max_row + 1
            ws1.cell(row=last, column=1, value="UKUPNO – SVI MATERIJALI").font = Font(bold=True)
            fill_green = PatternFill(start_color="EEF7EE", end_color="EEF7EE", fill_type="solid")
            headers = [c.value for c in ws1[1]]
            def idx(h): return headers.index(h) + 1 if h in headers else None
            c_kant = idx("Kant m"); c_m2 = idx("Površina m²"); c_rez = idx("Rezanje m")
            if c_kant: ws1.cell(row=last, column=c_kant, value=grand_total_kant).fill = fill_green
            if c_m2:   ws1.cell(row=last, column=c_m2,   value=grand_total_m2).fill = fill_green
            if c_rez:  ws1.cell(row=last, column=c_rez,  value=grand_total_rez).fill = fill_green
            fill_grey = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            for idx in subtotal_row_indices:
                excel_row = idx + 1
                for c in range(1, ws1.max_column+1):
                    cell = ws1.cell(row=excel_row+1, column=c)
                    cell.fill = fill_grey
                    if c in (1, c_kant or 0, c_m2 or 0, c_rez or 0):
                        cell.font = Font(bold=True)

        # ========== Sheet 2: Sažetak ==========
        ws2_name = "Sažetak"
        sum_by_traka.to_excel(writer, index=False, sheet_name=ws2_name)
        ws2 = writer.sheets[ws2_name]
        startrow = len(sum_by_traka) + 2
        extra = pd.DataFrame({
            "Traka": [
                "— Korpusi Iveral m²",
                "— Fronte Iveral m²",
                "— Leđa HDF m²",
                "— Rezanje m ukupno",
                "— Zbroj materijala m² ukupno",
            ],
            "Kant m ukupno": [
                korp_iveral_m2, fronte_iveral_m2, hdf_m2, total_rez_m,
                total_materials_m2,
            ],
        })
        extra.to_excel(writer, index=False, sheet_name=ws2_name, startrow=startrow)
        if engine == "xlsxwriter":
            ws2.set_column(0, 0, 30); ws2.set_column(1, 1, 24)

        # ========== Sheet 3: Narudžba ==========
        narudzba_cols = ["Oznaka","Naziv","Mat","Traka","A (mm)","B (mm)","Kom","Oznaka kantiranja"]
        df_n = df[narudzba_cols].copy()
        df_n.to_excel(writer, index=False, sheet_name="Narudžba")
        ws3 = writer.sheets["Narudžba"]
        if engine == "xlsxwriter":
            left = writer.book.add_format({"align":"left"})
            center = writer.book.add_format({"align":"center"})
            ws3.set_column(0, 0, 10, center)
            ws3.set_column(1, 1, 28, left)
            ws3.set_column(2, 3, 18, center)
            ws3.set_column(4, 5, 12, center)
            ws3.set_column(6, 6, 8,  center)
            ws3.set_column(7, 7, 16, center)
            ws3.autofilter(0, 0, len(df_n), len(narudzba_cols)-1)
            ws3.freeze_panes(1, 1)
        else:
            from openpyxl.styles import Alignment, Font
            ws3.freeze_panes = "B2"; ws3.auto_filter.ref = ws3.dimensions
            for cell in ws3[1]:
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
            for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
                if row and row[1].value is not None:
                    row[1].alignment = Alignment(horizontal="left")
                for c in [0,2,3,4,5,6,7]:
                    row[c].alignment = Alignment(horizontal="center")

        # ========== Sheet 4: Okov_Oprema_Dodatci ==========
        rows_e = []
        for r in okov_rows + oprema_rows:
            rows_e.append({
                "Kategorija": r["kategorija"], "Art. Nr.": r["art_nr"], "Naziv": r["naziv"],
                "Dobavljač": r["dobavljac"], "Jedinica": r["jedinica"],
                "Cijena (€)": r["cijena_eur"], "Količina": r["kolicina"], "Iznos (€)": r["iznos"]
            })
        for r in dodatci_rows:
            rows_e.append({
                "Kategorija": "DODATAK", "Šifra/Dod": r.get("sifra",""),
                "Naziv": r["naziv"], "Jedinica": r["jedinica"], "Vrsta": r["vrsta"],
                "A (mm)": r["A_mm"], "B (mm)": r["B_mm"], "Kom": r["kom"],
                "Obračun količina": r["obračun_količina"], "Cijena (€)": r["cijena_eur"],
                "Iznos (€)": r["iznos"]
            })

        df_e = pd.DataFrame(rows_e)
        if df_e.empty:
            df_e = pd.DataFrame([{"Kategorija":"","Art. Nr.":"","Naziv":"","Dobavljač":"","Jedinica":"","Cijena (€)":"","Količina":"","Iznos (€)":""}])
        df_e.to_excel(writer, index=False, sheet_name="Okov_Oprema_Dodatci")
        ws4 = writer.sheets["Okov_Oprema_Dodatci"]

        if engine == "xlsxwriter":
            center = writer.book.add_format({"align":"center"})
            left = writer.book.add_format({"align":"left"})
            ws4.autofilter(0, 0, len(df_e), len(df_e.columns)-1)
            ws4.freeze_panes(1, 1)
            ws4.set_column(0, len(df_e.columns)-1, 14, center)
            # "Naziv" i "Dobavljač" lijevo + šire
            try:
                naziv_idx = list(df_e.columns).index("Naziv")
                ws4.set_column(naziv_idx, naziv_idx, 34, left)
            except ValueError:
                pass
            try:
                dob_idx = list(df_e.columns).index("Dobavljač")
                ws4.set_column(dob_idx, dob_idx, 22, left)
            except ValueError:
                pass
        else:
            from openpyxl.styles import Alignment, Font
            ws4.freeze_panes = "B2"; ws4.auto_filter.ref = ws4.dimensions
            for cell in ws4[1]:
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
            headers = [c.value for c in ws4[1]]
            for col_name, align in [("Naziv","left"), ("Dobavljač","left")]:
                if col_name in headers:
                    col = headers.index(col_name)+1
                    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
                        row[col-1].alignment = Alignment(horizontal=align)

    bio.seek(0)
    return bio.getvalue(), None

# =============== RUN: Izračun + Izvoz ===============
st.markdown('<div class="sticky"></div>', unsafe_allow_html=True)
if st.button("🧮 Izračunaj ▶", use_container_width=True):

    # 1) Izvedi elemente
    rows = derive_rows(
        W,H,D,t,n_police,
        include_back, default_mat, default_traka,
        pod_vrsta_vanjski, kapa_vrsta_vanjska,
        include_kapa_povez, kapa_povez_mode, kapa_povez_sirina_mm, kapa_povez_posto,
        include_fronta, fronta_tip, fronta_montaza,
        razmak_hor, razmak_ver, razmak_srednji,
        preklop_hor, preklop_ver, default_mat_fr, default_traka_fr,
        include_haupt_hor, include_haupt_ver, haupt_sirina_mm
    )

    st.markdown("### 6) 📋 Sažetak elemenata i troškovnik")

    # 2) PRIKAZ u editoru (centriranje svih osim Naziv)
    display_rows = []
    for r in rows:
        d = dict(r)
        d["oznaka"] = short_code_for(d.get("naziv",""))
        d["mat"] = MAT_LABEL.get(d.get("mat", ""), d.get("mat", ""))
        d["traka"] = TRAK_LABEL.get(d.get("traka", ""), d.get("traka", ""))
        A = float(d.get("A_mm", 0) or 0); B = float(d.get("B_mm", 0) or 0)
        dugi, kratki = auto_kant_counts(
            naziv=str(d.get("naziv","")),
            auto=bool(d.get("auto", False)),
            A=A, B=B,
            fallback_dugi=int(d.get("kant_dugi", 0) or 0),
            fallback_kratki=int(d.get("kant_kratki", 0) or 0),
        )
        d["kant_dugi"] = dugi; d["kant_kratki"] = kratki
        display_rows.append(d)

    st.markdown("""
    <style>
    .center-editor [data-testid="stDataEditor"] thead th,
    .center-editor [data-testid="stDataEditor"] tbody td { text-align: center !important; }
    .center-editor [data-testid="stDataEditor"] thead th:nth-child(1),
    .center-editor [data-testid="stDataEditor"] tbody td:nth-child(1) { text-align: left !important; }
    </style>
    """, unsafe_allow_html=True)
    st.markdown('<div class="center-editor">', unsafe_allow_html=True)

    column_order = ["naziv","oznaka","mat","traka","A_mm","B_mm","kom","kant_dugi","kant_kratki","auto"]
    edited = st.data_editor(
        display_rows,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "naziv": st.column_config.TextColumn("Naziv"),
            "oznaka": st.column_config.TextColumn("Oznaka"),
            "A_mm": st.column_config.NumberColumn("Dim A (mm)", min_value=1, step=1),
            "B_mm": st.column_config.NumberColumn("Dim B (mm)", min_value=1, step=1),
            "kom": st.column_config.NumberColumn("Kom", min_value=1, step=1),
            "kant_dugi": st.column_config.SelectboxColumn("Kant DUGI", options=[0, 1, 2]),
            "kant_kratki": st.column_config.SelectboxColumn("Kant KRATKI", options=[0, 1, 2]),
            "auto": st.column_config.CheckboxColumn("✔️ Auto pravilo"),
            "mat": st.column_config.SelectboxColumn("Materijal", options=sorted(set(MAT_LABEL.values()))),
            "traka": st.column_config.SelectboxColumn("ABS traka", options=sorted(set(TRAK_LABEL.values()))),
        },
        column_order=column_order,
        key="editor_korpus",  # jedinstveni ključ
    )
    st.markdown('</div>', unsafe_allow_html=True)

    # 3) Normalizacija label -> šifra
    normalized_rows = []
    for r in edited:
        r = dict(r)
        if r.get("mat") in MAT_BY_LABEL:   r["mat"] = MAT_BY_LABEL[r["mat"]]
        if r.get("traka") in TRAK_BY_LABEL:r["traka"] = TRAK_BY_LABEL[r["traka"]]
        normalized_rows.append(r)

    # 4) Izračun – korpus
    report, metrics = calculate(normalized_rows, rez_usl, kant_usl, MATS, TRAK, FRONTS, FTRAK, USLG)
    mats_services_total, _ = materials_services_summary(metrics, use_waste, waste_pct)

    # 5) Izračun – okov/oprema/dodatci
    extras_total_val = extras_totals(okov_rows, oprema_rows, dodatci_rows)

    # 6) Rad i završni zbir
    labor_total_val = labor_total_calc(h_tp, r_tp, h_cnc, r_cnc, h_skl, r_skl, h_pak, r_pak)
    ukupno = final_summary_grand(mats_services_total, extras_total_val, labor_total_val, use_markup, markup_pct)

    st.success(f"✅ UKUPNO: {fmt_eur(ukupno)}")

    # CSV export (korpus elementi)
    st.markdown("### 📤 Izvoz")
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"izracun_unified_{timestamp}.csv"
    csv_buffer = io.StringIO()
    if report:
        fieldnames = list(report[0].keys())
        writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
        writer.writeheader()
        for row in report:
            writer.writerow(row)
    st.download_button(
        "⬇️ CSV – elementi (korpus)",
        data=csv_buffer.getvalue().encode("utf-8"),
        file_name=csv_filename,
        mime="text/csv",
        use_container_width=True
    )

    # --- XLSX export: komplet ---
    xlsx_bytes, xlsx_err = build_xlsx_kantiranje(report, normalized_rows, okov_rows, oprema_rows, dodatci_rows)
    if xlsx_err:
        st.error(f"XLSX izvoz nije uspio: {xlsx_err}")
    else:
        st.download_button(
            "⬇️ XLSX – komplet (korpus + narudžba + okov/oprema/dodatci)",
            data=xlsx_bytes,
            file_name=f"kantiranje_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    # --- PDF export s NAZIV naslovom (opcionalno) ---
    try:
        pdf_bytes = build_full_pdf(
            report, metrics, mats_services_total, extras_total_val, labor_total_val, use_markup, markup_pct,
            W, H, D, n_police, waste_pct, rok_dani,
            include_back, pod_vrsta_vanjski, kapa_vrsta_vanjska, include_kapa_povez,
            include_fronta, fronta_montaza=fronta_montaza, fronta_tip=fronta_tip
        )
        st.download_button(
            "⬇️ PDF – ponuda (s NAZIV naslovom)",
            data=pdf_bytes,
            file_name=f"ponuda_{timestamp}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
    except Exception as e:
        st.warning(f"PDF nije generiran: {e}")

else:
    st.info("Popunite korake 1–5, pa kliknite **🧮 Izračunaj ▶**.")

# =============== Loader cache reset ===============
if st.sidebar.button("🔄 Učitaj ponovno cjenik"):
    load_local.clear()
    load_from_uploaded.clear()
    load_from_csv_urls.clear()
    st.experimental_rerun()
